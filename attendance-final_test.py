#!/usr/bin/env python
# coding: utf-8

# In[ ]:


import openpyxl
import datetime
time = datetime.datetime.now()
cur_time = time.strftime("%H:%M")
cur_date = time.strftime("%d|%m|%y")

#print("Enter excel file name  [e.g sample.xlsx]  ",end="")
#file_name = input()
file_name="sample2.xlsx"
print("Enter current Date  ['01|02|21 format]  ",end="")
Date = input()

def createNewSheet(file_name,date ="01|01|21", sheetDate=cur_date): 
    #sheet Name is by defalut current date name
    global wb
    wb = openpyxl.load_workbook(f"{file_name}")
    sheet = wb.sheetnames
    initial_sheet = wb[f"{sheet[0]}"]
    global row, col
    row = initial_sheet.max_row
    col = initial_sheet.max_column
    
    if cur_date == date:
        wb.create_sheet(sheetDate)  # creating new sheet
        global cur_sheet
        cur_sheet = wb.get_sheet_by_name(f"{sheetDate}")  # getting new sheet
        month = cur_date  # updating cur
        
        for i in range(1,row+1):
            if i == 1:
                for j in range(1, col+1):
                    cur_sheet.cell(i, j).value = initial_sheet.cell(i, j).value
            else:
                for j in range(1, 4):
                    cur_sheet.cell(i, j).value = initial_sheet.cell(i, j).value
                    
        
        return "new sheet  '"+ sheetDate + "'  created"
    else:
        return "Sheet already exists..."


def attendance_mark(label, sheetDate ):
    cur_sheet = wb[f'{sheetDate}']
    lec = 0
    #print(cur_sheet)
    for i in range(2, row+1):
        name = cur_sheet.cell(i,1).value 
        #print(name)
        if name.lower() == label.lower():
            print("label test")
            # ALL lecture between 8:30 to 3:30
            if cur_time >= "08:30" and cur_time <= "08:45":
                lec = 1  # first lecture
            elif cur_time >= "09:30" and cur_time <= "11:30":
                lec = 2  # second lecture
            elif cur_time >= "10:30" and cur_time <= "12:30":
                lec = 3  # third lecture
            elif cur_time >= "00:50" and cur_time <= "02:55":
                lec = 4  # fourth lecture
            elif cur_time >= "13:30" and cur_time <= "14:30":
                lec = 5  # fifth lecture
            elif cur_time >= "14:30" and cur_time <= "16:30":
                lec = 6  # sixth lecture
            attend = cur_sheet.cell(i,lec+3).value
            
            if str(attend) == "None":
                data = cur_sheet.cell(i, lec+3).value = cur_time
                data = cur_sheet.cell(i, 10).value = cur_date
                #return label + " marked"
            else:
                return ("Already marked !!!")
    
          
    wb.save('sample2.xlsx')
    
    

createNewSheet(file_name,Date)
label1 = "ajay kumar yadav"
sheetDate = cur_date
attendance_mark(label1,sheetDate)


# In[ ]:





# In[ ]:




