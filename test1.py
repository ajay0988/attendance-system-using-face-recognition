import cv2
import numpy as np
import face_recognition
import os
import time
import pyttsx3 as pt
import smtplib
import openpyxl
import datetime
time = datetime.datetime.now()
cur_time = time.strftime("%H:%M")
cur_date = time.strftime("%d|%m|%y")
file_name = "sample2.xlsx"

def createNewSheet(file_name, sheetDate=cur_date): 
    #sheet Name is by defalut current date name
    print("Enter current Date  ['01|02|21 format]  ",end="")
    date = input()
    wb = openpyxl.load_workbook(f"{file_name}")
    sheet = wb.sheetnames
    initial_sheet = wb[f"{sheet[0]}"]
    row = initial_sheet.max_row
    col = initial_sheet.max_column
    
    if cur_date == date:
        wb.create_sheet(sheetDate)  # creating new sheet
        cur_sheet = wb.get_sheet_by_name(f"{sheetDate}")  # getting new sheet
        month = cur_date  # updating cur
        
        for i in range(1,row+1):
            if i == 1:
                for j in range(1, col+1):
                    cur_sheet.cell(i, j).value = initial_sheet.cell(i, j).value
            else:
                for j in range(1, 4):
                    cur_sheet.cell(i, j).value = initial_sheet.cell(i, j).value
                cur_sheet.cell(i, 11).value = initial_sheet.cell(i, 11).value
                    
        wb.save('sample2.xlsx')
        pt.speak("new sheet is created")
        return "new sheet  '"+ sheetDate + "'  created"
        #pt.speak("new sheet is created")
    else:
        pt.speak("Sheet already exist")
        return "Sheet already exists..."
    

try:
    def sendMail( mail_to, content, my_mail="ar9131000@gmail.com", my_pass= "Ar054982@"):
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.ehlo()
        server.starttls()
        server.login(my_mail, my_pass) # eneter your email and password but you have to enable <less secure app> in your email privacy setting
        server.sendmail(my_mail, mail_to, content) # eneter your email, reciver email, content to send
        server.close()
except Exception as e:
    print("e")
    
def attendance_mark(label, sheetDate, file_name):
    wb = openpyxl.load_workbook(file_name)
    cur_sheet = wb[f'{sheetDate}']
    lec = 0
    row = cur_sheet.max_row
    col = cur_sheet.max_column
    #print(label.lower())
    #print(cur_sheet)
    for i in range(2, row+1):
        name = cur_sheet.cell(i,1).value 
        #print(name) 
        name = str(name)
        #print(type(name))
        if name.lower() == label.lower():
            #print("label test")
            # ALL lecture between 8:30 to 3:30
            if cur_time >= "08:30" and cur_time <= "08:45":
                lec = 1  # first lecture
            elif cur_time >= "08:45" and cur_time <= "11:30":
                lec = 2  # second lecture
            elif cur_time >= "10:30" and cur_time <= "12:30":
                lec = 3  # third lecture
            elif cur_time >= "12:30" and cur_time <= "13:30":
                lec = 4  # fourth lecture
            elif cur_time >= "13:30" and cur_time <= "14:30":
                lec = 5  # fifth lecture
            elif cur_time >= "14:30" and cur_time <= "18:30":
                lec = 6  # sixth lecture
            attend = cur_sheet.cell(i,lec+3).value
            
            if str(attend) == "None":
                cur_sheet.cell(i, lec+3).value = cur_time
                cur_sheet.cell(i, 10).value = cur_date
                mail = cur_sheet.cell(i, 11).value
                #print(mail)
                if str(mail) != "None":
                    content = name + " attend lecture " + str(lec)
                    sendMail(mail, content)
                    wb.save('sample2.xlsx')
                    pt.speak(f"{name} attendence marked")
                    return name +" Attendence 'marked' and 'mailed' to parent "
                #break
             
                else:
                    return name +"  parent mail not available"
            else:
                return ("Already marked !!!")
    
            
    
#label = "aryan raj"
#sheetDate = "20|03|21"
#attendance_mark(label,sheetDate, file_name)
#pt.speak("attendence done from our side")

os.system("cls")
pt.speak("My program now started")
path = 'train_image'
images = []
classNames = []
myList = os.listdir(path)
#print(myList)
for cl in myList:
    curImg = cv2.imread(f'{path}/{cl}')
    images.append(curImg)
    classNames.append(os.path.splitext(cl)[0])
print(classNames)

def findEncodings(images):
    encodeList = []


    for img in images:
        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        encode = face_recognition.face_encodings(img)[0]
        encodeList.append(encode)
    return encodeList



"""def markAttendance(name):
    with open('Attendance.csv', 'r+') as f:
        myDataList = f.readlines()


        nameList = []
        for line in myDataList:
            entry = line.split(',')
            nameList.append(entry[0])
            if name not in nameList:
                pt.speak("I will make attendance for you {}".format(name))
                now = datetime.now()
                dtString = now.strftime('%H:%M:%S')
                f.writelines(f'\n{name},{dtString}')
"""
#### FOR CAPTURING SCREEN RATHER THAN WEBCAM
# def captureScreen(bbox=(300,300,690+300,530+300)):
#     capScr = np.array(ImageGrab.grab(bbox))
#     capScr = cv2.cvtColor(capScr, cv2.COLOR_RGB2BGR)
#     return capScr
pt.speak("Process the data going on")
print("\n \t\tProcessing.....")
encodeListKnown = findEncodings(images)
print('\n\t\tEncoding Complete\n\n')
pt.speak("encode complete")
print("\t" ,end="")
#sheet_status = createNewSheet(file_name,Date)
sheet_status = createNewSheet(file_name, sheetDate=cur_date)
print(sheet_status)

print(sheet_status)
cap = cv2.VideoCapture(0)

while True:

    success, img = cap.read()
# img = captureScreen()
    imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
    imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)

    facesCurFrame = face_recognition.face_locations(imgS)
    encodesCurFrame = face_recognition.face_encodings(imgS, facesCurFrame)

    for encodeFace, faceLoc in zip(encodesCurFrame, facesCurFrame):
        matches = face_recognition.compare_faces(encodeListKnown, encodeFace)
        faceDis = face_recognition.face_distance(encodeListKnown, encodeFace)
# print(faceDis)
        matchIndex = np.argmin(faceDis)

        if matches[matchIndex]:
            #pt.speak("{} your face is detectd now".format(name))

            l_name = classNames[matchIndex].upper()
            #pt.speak("{} your face is detected ".format(l_name))
            #pt.speak(" I will make attendance for you {}".format(name))

            #time.sleep(3)
            
# print(name)
            y1, x2, y2, x1 = faceLoc
            y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
            cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 1)
            cv2.rectangle(img, (x1, y2 - 35), (x2, y2), (0, 255, 0), cv2.FILLED)
            cv2.putText(img, l_name, (x1 + 6, y2 - 6), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)
            #markAttendance(name)
            #label1 = "ajay kumar yadav"
            #sheetDate = cur_date
            #attendance_mark(l_name,sheetDate)
            #label = "aryan raj"
            #sheetDate = "20|03|21"
            attendance_mark(l_name,cur_date, file_name)
            #attendance_mark(l_name)


    cv2.imshow('Live Attendance Program', img)
    if cv2.waitKey(10)=='q':
        break
